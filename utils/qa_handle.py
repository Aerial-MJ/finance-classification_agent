import os
from openai import OpenAI
import openpyxl
from tqdm import tqdm
import math
import fitz
import markdown2
from openpyxl import load_workbook
import openpyxl
import argparse


def _parse_args():
    parser = argparse.ArgumentParser(
        description="handle questions based on a table"
    )
    parser.add_argument(
        "--question_table_path",
        type=str,
        default=r"Z:\wangyuxing\制度文件拆分问答对\2025\合作阿里微调项目\gene_q_project\questions_gene\questions_cunkuan.xlsx",
        help="The path of the question table.")

    parser.add_argument(
        "--out_path",
        type=str,
        default=r"Z:\wangyuxing\制度文件拆分问答对\2025\合作阿里微调项目\gene_q_project\questions_handle\模型直接生成\cunkuan/",
        help="The path of the question table.")

    args = parser.parse_args()

    return args



str2 = "问题："
str3 = "答案："
def question_handle(question_table_path, out_path):
    wb_cunkaun = openpyxl.load_workbook(question_table_path)
    print(wb_cunkaun.sheetnames)
    questions_cunkuan = wb_cunkaun['question']
    for i in range(questions_cunkuan.max_row-1):
        print(i+1)
        filename = questions_cunkuan.cell(row=i+2, column=1).value
        file_q_kehu = questions_cunkuan.cell(row=i+2, column=2).value
        if "</think>" not in file_q_kehu:
            print(filename)
        if filename is not None and file_q_kehu is not None and "</think>" in file_q_kehu:
            workbook_q_handle = openpyxl.Workbook()
            new_sheet1 = workbook_q_handle.create_sheet("question_handle",index=0)
            question_handle = workbook_q_handle["question_handle"]
            question_handle["A1"] = "问题"
            question_handle["B1"] = "答案"
            # print("这是文件名： ", filename)
            # print("---------------------------------------------")
            answer_kehu = file_q_kehu.split("</think>")[1]
            answer_list_kehu = answer_kehu.split("\n")
            j = 2
            for index, ans in enumerate(answer_list_kehu):
                if str2 in ans:
                    print(ans)
                    question_handle.cell(row=j, column=1).value = ans
                if str3 in ans:
                    print(ans)
                    question_handle.cell(row=j, column=2).value = ans
                    j += 1
            filename_table = filename + ".xlsx"
            workbook_q_handle.save(out_path + filename_table)
        print("===================================================")
#
args = _parse_args()
question_table_path = r"C:\Users\JNYH\Desktop\新文档\长文件\全部\qa2.xlsx"
out_path = r"C:\Users\JNYH\Desktop\新文档\长文件\qa-files-2/"
question_handle(question_table_path,out_path)

